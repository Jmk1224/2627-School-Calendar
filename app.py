import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

st.set_page_config(page_title="學校活動及排班管理系統", page_icon="📅", layout="wide")

st.title("🏫 學校活動及教師排班管理系統")
st.write("表格支援直接修改內容或刪除單筆資料。資料會同步更新於 Excel 與 Google Calendar。")

# Initialize Session State
if "activities" not in st.session_state:
    st.session_state.activities = [
        {
            "id": 1,
            "title": "[幼訪]協康會譚杜中心-家長訪校",
            "date": datetime.strptime("2026-09-05", "%Y-%m-%d").date(),
            "start_time": datetime.strptime("10:00", "%H:%M").time(),
            "end_time": datetime.strptime("12:00", "%H:%M").time(),
            "location": "禮堂",
            "teachers": "校",
            "notes": "備註說明範例",
            "exported": False
        }
    ]

# -------------------------------------------------------------------
# Sidebar: Add New Activity
# -------------------------------------------------------------------
st.sidebar.header("➕ 新增活動內容")
with st.sidebar.form("event_form", clear_on_submit=True):
    title = st.text_input("活動名稱 (例如: 家長日)")
    event_date = st.date_input("日期", value=datetime.today())
    col1, col2 = st.columns(2)
    with col1:
        start_time = st.time_input("開始時間", value=datetime.strptime("08:00", "%H:%M").time())
    with col2:
        end_time = st.time_input("結束時間", value=datetime.strptime("12:00", "%H:%M").time())
    location = st.text_input("地點 (例如: 禮堂)")
    teachers_raw = st.text_input("負責老師", placeholder="例如: 明、陳、李")
    notes = st.text_input("說明 / 備註", placeholder="例如: 需準備投影設備")

    submitted = st.form_submit_button("新增活動")
    if submitted:
        if title and teachers_raw:
            new_id = max([a["id"] for a in st.session_state.activities], default=0) + 1
            st.session_state.activities.append({
                "id": new_id,
                "title": title,
                "date": event_date,
                "start_time": start_time,
                "end_time": end_time,
                "location": location,
                "teachers": teachers_raw.replace(" ", "、").replace(",", "、"),
                "notes": notes,
                "exported": False
            })
            st.sidebar.success(f"已新增：{title}")
            st.rerun()
        else:
            st.sidebar.error("請填寫活動名稱及負責老師！")

# -------------------------------------------------------------------
# Main Content: Editable Data Table
# -------------------------------------------------------------------
st.subheader("📋 所有歷史及新增活動 (可直接在表格內修改或刪除)")

if st.session_state.activities:
    # Convert list to DataFrame for editable view
    df = pd.DataFrame([
        {
            "id": act["id"],
            "活動名稱": act["title"],
            "日期": act["date"].strftime("%Y-%m-%d"),
            "開始時間": act["start_time"].strftime("%H:%M"),
            "結束時間": act["end_time"].strftime("%H:%M"),
            "地點": act["location"],
            "負責老師": act["teachers"],
            "說明": act["notes"],
            "Google Calendar 狀態": "✅ 已匯入" if act["exported"] else "🆕 新增 (待匯入)"
        } for act in st.session_state.activities
    ])

    # Display Editable Data Table
    edited_df = st.data_editor(
        df,
        num_rows="dynamic",  # Allows deleting rows
        use_container_width=True,
        hide_index=True,
        column_config={
            "id": st.column_config.NumberColumn("序號", disabled=True),
            "Google Calendar 狀態": st.column_config.TextColumn("Google Calendar 狀態", disabled=True)
        },
        key="data_editor"
    )

    # Sync back changes from Table back to Session State
    if st.button("💾 儲存表格修改"):
        new_activities = []
        for _, row in edited_df.iterrows():
            # Find original export status
            orig = next((item for item in st.session_state.activities if item["id"] == row["id"]), None)
            exported_status = orig["exported"] if orig else False

            try:
                parsed_date = datetime.strptime(str(row["日期"]), "%Y-%m-%d").date()
                parsed_start = datetime.strptime(str(row["開始時間"]), "%H:%M").time()
                parsed_end = datetime.strptime(str(row["結束時間"]), "%H:%M").time()
            except ValueError:
                st.error(f"日期或時間格式不正確，日期請用 YYYY-MM-DD，時間請用 HH:MM。")
                st.stop()

            new_activities.append({
                "id": int(row["id"]),
                "title": str(row["活動名稱"]),
                "date": parsed_date,
                "start_time": parsed_start,
                "end_time": parsed_end,
                "location": str(row["地點"]),
                "teachers": str(row["負責老師"]),
                "notes": str(row["說明"]),
                "exported": exported_status
            })
        st.session_state.activities = new_activities
        st.success("修改已成功儲存！")
        st.rerun()

    if st.button("🗑️ 清空所有數據"):
        st.session_state.activities = []
        st.rerun()

    st.markdown("---")
    st.subheader("📥 下載專區")

    # ---------------------------------------------------------------
    # Excel Builder
    # ---------------------------------------------------------------
    def build_excel(data_list):
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="微軟正黑體", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        # Sheet 1: 所有活動總表
        ws1 = wb.active
        ws1.title = "所有活動總表"
        ws1.append(["序號", "活動名稱", "日期", "時間", "地點", "負責老師", "說明", "Calendar 狀態"])

        for item in data_list:
            teachers_list = [t.strip() for t in item["teachers"].split("、") if t.strip()]
            formatted_date = item["date"].strftime("%d/%m/%Y")
            time_str = f"{item['start_time'].strftime('%H:%M')} - {item['end_time'].strftime('%H:%M')}"
            status = "已匯入" if item["exported"] else "未匯入"
            ws1.append([item["id"], item["title"], formatted_date, time_str, item["location"], "、".join(teachers_list), item["notes"], status])

        for row in ws1.iter_rows(min_row=1, max_row=len(data_list)+1, min_col=1, max_col=8):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # Sheet 2: 老師分類表 (矩陣)
        ws2 = wb.create_sheet(title="老師分類表")
        all_teachers = sorted(list(set(t.strip() for item in data_list for t in item["teachers"].split("、") if t.strip())))
        all_dates = sorted(list(set(item["date"] for item in data_list)))

        ws2.append(["日期"] + all_teachers)
        grid_data = {}
        for item in data_list:
            d = item["date"]
            teachers = [t.strip() for t in item["teachers"].split("、") if t.strip()]
            for t in teachers:
                if (d, t) in grid_data:
                    grid_data[(d, t)] += f"、{item['title']}"
                else:
                    grid_data[(d, t)] = item["title"]

        for d in all_dates:
            row_data = [d.strftime("%d/%m/%Y")]
            for t in all_teachers:
                row_data.append(grid_data.get((d, t), ""))
            ws2.append(row_data)

        for row in ws2.iter_rows(min_row=1, max_row=len(all_dates)+1, min_col=1, max_col=len(all_teachers)+1):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ---------------------------------------------------------------
    # ICS Builder
    # ---------------------------------------------------------------
    def build_ics(data_list, export_only_new=True):
        ics_content = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//School Executive Secretary//NONSGML Event//EN"
        ]
        
        target_items = [item for item in data_list if not item["exported"]] if export_only_new else data_list

        for item in target_items:
            teachers_list = [t.strip() for t in item["teachers"].split("、") if t.strip()]
            teachers_str = "、".join(teachers_list)
            summary = f"{item['title']}（{teachers_str}）"
            desc = f"負責老師：{teachers_str}\\n說明：{item['notes']}"

            start_dt = f"{item['date'].strftime('%Y%m%d')}T{item['start_time'].strftime('%H%M%S')}"
            end_dt = f"{item['date'].strftime('%Y%m%d')}T{item['end_time'].strftime('%H%M%S')}"

            ics_content.extend([
                "BEGIN:VEVENT",
                f"SUMMARY:{summary}",
                f"LOCATION:{item['location']}",
                f"DESCRIPTION:{desc}",
                f"DTSTART;TZID=Asia/Hong_Kong:{start_dt}",
                f"DTEND;TZID=Asia/Hong_Kong:{end_dt}",
                "END:VEVENT"
            ])
        ics_content.append("END:VCALENDAR")
        return "\n".join(ics_content), len(target_items)

    col_a, col_b = st.columns(2)
    
    with col_a:
        excel_bytes = build_excel(st.session_state.activities)
        st.download_button(
            label="📊 下載完整 Excel 報表",
            data=excel_bytes,
            file_name="學校活動與教師排班表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    with col_b:
        ics_text, new_count = build_ics(st.session_state.activities, export_only_new=True)
        if new_count > 0:
            if st.download_button(
                label=f"📅 下載未匯入的 Google 日曆檔 ({new_count} 個新活動)",
                data=ics_text,
                file_name="new_school_events.ics",
                mime="text/calendar"
            ):
                for act in st.session_state.activities:
                    act["exported"] = True
                st.success("已更新匯入狀態！")
        else:
            st.info("目前所有活動皆已匯入 Google Calendar，無新增項目。")

else:
    st.info("目前沒有任何活動紀錄。")
